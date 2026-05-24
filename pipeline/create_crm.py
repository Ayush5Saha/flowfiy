import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Investor Pipeline"

# Column headers
headers = [
    "Date Added", "Name", "Company", "Role", "Country", "City",
    "Website", "Email", "LinkedIn", "Lead Source", "Qual Score",
    "Gap Identified", "Outreach Channel", "Email Status", "DM Status",
    "Follow-up Stage", "Reply Status", "Meeting Booked", "Meeting Date", "Notes"
]

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
alt_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

today = str(date.today())

# Investor data — all 20 investors
investors = [
    {
        "name": "Vaibhav Domkundwar",
        "company": "Better Capital",
        "role": "Founder & General Partner",
        "country": "India",
        "city": "Bengaluru",
        "website": "bettercapital.us",
        "email": "vaibhav@bettercapital.us",
        "linkedin": "linkedin.com/in/vdomkundwar",
        "source": "Vibe Prospecting",
        "score": 9,
        "gap": "Pre-seed/seed AI SaaS; invested in dev tools and AI startups",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "Sent",
        "notes": "DM sent prev session. Better Capital focus: pre-seed Indian startups."
    },
    {
        "name": "Nitin Sharma",
        "company": "Antler India",
        "role": "Partner",
        "country": "India",
        "city": "Bengaluru",
        "website": "antler.co",
        "email": "nitin@antler.co",
        "linkedin": "linkedin.com/in/nitinsharma",
        "source": "Vibe Prospecting",
        "score": 8,
        "gap": "Antler backs early-stage AI/SaaS; Nitin leads India portfolio",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "Sent",
        "notes": "1st-degree connection. DM sent successfully."
    },
    {
        "name": "Raghav Goyal",
        "company": "Antler",
        "role": "Venture Partner",
        "country": "India",
        "city": "Bengaluru",
        "website": "antler.co",
        "email": "raghav@antler.co",
        "linkedin": "linkedin.com/in/raghavgoyal",
        "source": "Vibe Prospecting",
        "score": 8,
        "gap": "Operator-angel; deep B2B SaaS experience; Antler portfolio",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "Sent",
        "notes": "Open profile. DM sent successfully."
    },
    {
        "name": "Karan Mohla",
        "company": "B Capital",
        "role": "General Partner",
        "country": "India",
        "city": "Bengaluru",
        "website": "bcapital.com",
        "email": "karan@bcapital.com",
        "linkedin": "linkedin.com/in/karanmohla",
        "source": "Vibe Prospecting",
        "score": 9,
        "gap": "B Capital backs B2B SaaS, AI infra; Karan leads India/SEA deals",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "Sent",
        "notes": "Open profile. DM sent successfully."
    },
    {
        "name": "Tony Wang",
        "company": "500 Global",
        "role": "Partner",
        "country": "Singapore",
        "city": "Singapore",
        "website": "500.co",
        "email": "tony@500.co",
        "linkedin": "linkedin.com/in/tonywang500",
        "source": "Vibe Prospecting",
        "score": 8,
        "gap": "500 Global SEA/India focus; invests in AI productivity tools",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "Sent",
        "notes": "Open profile. DM sent successfully."
    },
    {
        "name": "Trevor Hazlett",
        "company": "Entrepreneur First",
        "role": "Investment Manager",
        "country": "UK",
        "city": "London",
        "website": "joinef.com",
        "email": "trevor@joinef.com",
        "linkedin": "linkedin.com/in/trevorhazlett",
        "source": "Vibe Prospecting",
        "score": 7,
        "gap": "EF backs technical founders in AI/SaaS at pre-seed",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "Sent",
        "notes": "Open profile. DM sent successfully."
    },
    {
        "name": "Rajan Anandan",
        "company": "Peak XV Partners (Sequoia India)",
        "role": "Managing Director",
        "country": "India",
        "city": "Bengaluru",
        "website": "peakxv.com",
        "email": "rajan@sequoiacap.com",
        "linkedin": "linkedin.com/in/rajananandan",
        "source": "Vibe Prospecting",
        "score": 9,
        "gap": "Peak XV backs AI SaaS; Rajan is top India tech investor",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "InMail Required",
        "notes": "Profile requires LinkedIn Premium InMail. Email sent."
    },
    {
        "name": "Pete Koomen",
        "company": "Y Combinator",
        "role": "Group Partner",
        "country": "USA",
        "city": "San Francisco",
        "website": "ycombinator.com",
        "email": "pete@ycombinator.com",
        "linkedin": "linkedin.com/in/petekoomen",
        "source": "Manual Research",
        "score": 9,
        "gap": "YC GP; invested in Optimizely; understands AI dev tools deeply",
        "channel": "Email",
        "email_status": "Sent",
        "dm_status": "No Message Button",
        "notes": "No LinkedIn message button available. Email sent."
    },
    {
        "name": "Shane Shin",
        "company": "Shorooq Partners",
        "role": "Founding Partner",
        "country": "UAE",
        "city": "Abu Dhabi",
        "website": "shorooq.com",
        "email": "shane@shorooq.com",
        "linkedin": "linkedin.com/in/shaneykshin",
        "source": "Vibe Prospecting",
        "score": 9,
        "gap": "Shorooq is top MENA VC; Shane invests in AI/SaaS for MENA market",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "InMail Required",
        "notes": "Correct profile: shaneykshin. InMail required. Email sent."
    },
    {
        "name": "Akash Venkat",
        "company": "Entrepreneur First",
        "role": "Investor",
        "country": "India",
        "city": "Bengaluru",
        "website": "joinef.com",
        "email": "akash@joinef.com",
        "linkedin": "linkedin.com/in/akashvenkat",
        "source": "Vibe Prospecting",
        "score": 7,
        "gap": "EF India; pre-seed AI/SaaS focus",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "InMail Required",
        "notes": "2nd-degree but no Open Profile. InMail required. Email sent."
    },
    {
        "name": "Ankur Gupta",
        "company": "Accel India",
        "role": "Partner",
        "country": "India",
        "city": "Bengaluru",
        "website": "accel.com",
        "email": "ankurg@accel.com",
        "linkedin": "linkedin.com/in/ankurgupta",
        "source": "Vibe Prospecting",
        "score": 8,
        "gap": "Accel backs enterprise SaaS and AI tools; Ankur focuses India/SEA",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "InMail Required",
        "notes": "InMail required. Email sent."
    },
    {
        "name": "Girish Mathrubootham",
        "company": "Together Fund",
        "role": "Founder & General Partner",
        "country": "India",
        "city": "Chennai",
        "website": "together.fund",
        "email": "girish@together.fund",
        "linkedin": "linkedin.com/in/girishmathrubootham",
        "source": "Manual Research",
        "score": 9,
        "gap": "Freshworks founder turned VC; deep B2B SaaS + AI operator-angel",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "InMail Required",
        "notes": "High priority operator-angel. InMail required. Email sent."
    },
    {
        "name": "Marion Bazille",
        "company": "Big Idea Ventures",
        "role": "Partner",
        "country": "Singapore",
        "city": "Singapore",
        "website": "bigideaventures.com",
        "email": "marion@bigideaventures.com",
        "linkedin": "linkedin.com/in/marionbazille",
        "source": "Vibe Prospecting",
        "score": 7,
        "gap": "BIV invests in AI-enabled B2B; Marion leads SEA portfolio",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "InMail Required",
        "notes": "InMail required. Email sent."
    },
    {
        "name": "Hetvi Lalan",
        "company": "Venture Catalysts++",
        "role": "Principal",
        "country": "India",
        "city": "Mumbai",
        "website": "vcatalysts.co",
        "email": "hetvi@vcatalysts.co",
        "linkedin": "linkedin.com/in/hetvilalan",
        "source": "Vibe Prospecting",
        "score": 7,
        "gap": "VCats++ backs Indian B2B SaaS and AI startups at seed",
        "channel": "Email + LinkedIn DM",
        "email_status": "Sent",
        "dm_status": "InMail Required",
        "notes": "InMail required. Email sent."
    },
    {
        "name": "Zak Doric",
        "company": "a16z",
        "role": "Partner",
        "country": "USA",
        "city": "San Francisco",
        "website": "a16z.com",
        "email": "zak@a16z.com",
        "linkedin": "linkedin.com/in/zakdoric",
        "source": "Manual Research",
        "score": 6,
        "gap": "a16z AI fund; seed checks in AI infrastructure and developer tools",
        "channel": "Email",
        "email_status": "Sent",
        "dm_status": "3rd+ Degree",
        "notes": "3rd+ degree connection on LinkedIn. Email only."
    },
    {
        "name": "Ronny Shibley",
        "company": "Antler UK",
        "role": "Partner",
        "country": "UK",
        "city": "London",
        "website": "antler.co",
        "email": "ronny@antler.co",
        "linkedin": "linkedin.com/in/ronnyshibley",
        "source": "Vibe Prospecting",
        "score": 7,
        "gap": "Antler UK pre-seed AI/SaaS; backs technical founders",
        "channel": "Email",
        "email_status": "Sent",
        "dm_status": "3rd+ Degree",
        "notes": "3rd+ degree. Email only."
    },
    {
        "name": "Jessica Anderson",
        "company": "Techstars",
        "role": "Managing Director",
        "country": "USA",
        "city": "New York",
        "website": "techstars.com",
        "email": "jessica@techstars.com",
        "linkedin": "linkedin.com/in/jessicaanderson",
        "source": "Vibe Prospecting",
        "score": 7,
        "gap": "Techstars accelerator; backs AI SaaS and B2B tools globally",
        "channel": "Email",
        "email_status": "Sent",
        "dm_status": "3rd+ Degree",
        "notes": "3rd+ degree. Email only."
    },
    {
        "name": "Jeff Giles",
        "company": "Garden City Equity",
        "role": "General Partner",
        "country": "USA",
        "city": "Atlanta",
        "website": "gardencityequity.com",
        "email": "jeff@gardencityequity.com",
        "linkedin": "linkedin.com/in/jeffgiles",
        "source": "Manual Research",
        "score": 6,
        "gap": "Invests in AI SaaS and B2B software; angel + seed focus",
        "channel": "Email",
        "email_status": "Sent",
        "dm_status": "3rd+ Degree",
        "notes": "3rd+ degree. Email only."
    },
    {
        "name": "Xueyuan Wang",
        "company": "Sequoia Capital",
        "role": "Partner",
        "country": "USA",
        "city": "Menlo Park",
        "website": "sequoiacap.com",
        "email": "xueyuan@sequoiacap.com",
        "linkedin": "linkedin.com/in/xueyuanwang",
        "source": "Manual Research",
        "score": 6,
        "gap": "Sequoia scout for AI/SaaS; seed to Series A",
        "channel": "Email",
        "email_status": "Sent",
        "dm_status": "3rd+ Degree / Wrong Profile",
        "notes": "Could not confirm correct LinkedIn profile. Email only."
    },
    {
        "name": "Katherine Boyle",
        "company": "a16z",
        "role": "General Partner",
        "country": "USA",
        "city": "San Francisco",
        "website": "a16z.com",
        "email": "kboyle@a16z.com",
        "linkedin": "Not Found",
        "source": "Manual Research",
        "score": 7,
        "gap": "a16z GP; backs American dynamism + AI tools; seed/Series A",
        "channel": "Email",
        "email_status": "Sent",
        "dm_status": "Profile Not Found",
        "notes": "Correct LinkedIn profile could not be found. Email sent only."
    },
]

# Status color map
status_colors = {
    "Sent": "C6EFCE",          # green
    "InMail Required": "FFEB9C",  # yellow
    "3rd+ Degree": "FFEB9C",
    "3rd+ Degree / Wrong Profile": "FFEB9C",
    "No Message Button": "FFEB9C",
    "Profile Not Found": "FFD7D7",  # light red
    "Not Sent": "FFD7D7",
}

# Write data rows
for row_idx, inv in enumerate(investors, 2):
    row_data = [
        today,
        inv["name"],
        inv["company"],
        inv["role"],
        inv["country"],
        inv["city"],
        inv["website"],
        inv["email"],
        inv["linkedin"],
        inv["source"],
        inv["score"],
        inv["gap"],
        inv["channel"],
        inv["email_status"],
        inv["dm_status"],
        "None",
        "No Reply",
        "No",
        "",
        inv["notes"],
    ]

    fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = left_align if col_idx not in (1, 10, 11, 13, 14, 15, 16, 17, 18) else center_align

        # Apply status color to Email Status (col 14) and DM Status (col 15)
        if col_idx == 14:
            color = status_colors.get(str(value), "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        elif col_idx == 15:
            color = status_colors.get(str(value), "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        elif col_idx == 11:
            # Score — color code
            if isinstance(value, int):
                if value >= 9:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True, color="375623")
                elif value >= 7:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(bold=True, color="9C5700")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            cell.fill = fill

# Column widths
col_widths = [12, 22, 28, 24, 10, 14, 28, 32, 36, 16, 8, 45, 16, 12, 22, 14, 12, 14, 12, 48]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze top row
ws.freeze_panes = "A2"

# Row height for header
ws.row_dimensions[1].height = 30

import os
os.makedirs("pipeline", exist_ok=True)
wb.save("pipeline/ayush-outbound-pipeline.xlsx")
print("Excel CRM created successfully: pipeline/ayush-outbound-pipeline.xlsx")
