# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

headers = ["Lead No.","Company Name","Website","Country","City","Company Type","Product Category",
"Relevant Products","Official Email","Official Phone Number","WhatsApp Number","WhatsApp Verification Status",
"Contact Page URL","Impressum Page URL","LinkedIn Company Page","Source URL","Recent Activity Proof",
"Buyer Fit Score /10","Reason for Score","Best Outreach Angle"]

leads = [
[1,"Hermann Schulz GmbH (Hamburger Gewuerz-Muehle / HGM Spice)","http://www.hgmspice.de","Germany","Hamburg",
"Spice manufacturer + direct importer + wholesaler/exporter (founded 1930)",
"Spices A-Z · Blends & Salts · Bio (Organic) spices · Specialty items",
"Single-origin raw spices (turmeric, chilli, cumin, coriander, pepper etc.), spice blends, organic spices — direct global import & in-house milling (14 mills, 8 cleaning + 6 mixing systems)",
"info@hgmspice.de","+49 40 7897010","Not publicly available","Not publicly verified",
"http://www.hgmspice.de/kontakt/","http://www.hgmspice.de/kontakt/impressum/","Not publicly available",
"https://www.google.com/maps/search/?api=1&query=Hamburger%20Gew%C3%BCrz-M%C3%BChle%20Hermann%20Schulz%20GmbH&query_place_id=ChIJQ31jfi6MsUcRufZHhuo6Ia8",
"Active 2026 — 90+ yrs (founded 1930); FSSC 22000 certified; operates 14 mills / 8 cleaning / 6 mixing lines; MD Holger Schulz named via Impressum; HRB 30714 Hamburg; 24 Google reviews (4.3)",
9,"Direct global spice importer + miller = buys exactly Tbaisa's raw material; named MD (Holger Schulz) via Impressum; FSSC 22000 means they need spec-compliant, ETO-free origin supply",
"Email info@hgmspice.de to Herr Holger Schulz as GF; offer India-direct ETO-free, EU-MRL-compliant raw spices ex-Mundra for their milling lines + Bio (organic) range; attach Tbaisa spec sheet"],

[2,"Shalimar Foods GmbH","https://shalimar-food.com","Germany","Cologne (Koeln)",
"B2B food wholesaler + cash & carry + daily delivery (est. ~2001)",
"Indian Specialties · International (Italian/Mexican/Chinese) · Gastronomy supply",
"1200+ SKUs incl. explicit 'Indische Spezialitaeten' line — spices, lentils/pulses, basmati rice, ethnic foods for gastronomy & retail trade",
"info@shalimar-foods.com","+49 2203 459690","Not publicly available","Not publicly verified",
"https://shalimar-food.com/kontakt/","https://shalimar-food.com/impressum/","Not publicly available",
"https://www.google.com/maps/search/?api=1&query=Shalimar%20Foods%20GmbH&query_place_id=ChIJz9g2eQ0nv0cRCtAlpWGQx7E",
"Active 2026 — 23+ yrs trading; 1200+ products; cash & carry pickup market + daily delivery service; 24 Google reviews (4.8)",
9,"South-Asian specialist wholesaler with an explicit Indian-products line and gastronomy distribution = natural buyer of India-direct spices/pulses/basmati; strong rating shows live, active trade",
"Email info@shalimar-foods.com pitching India-direct spices + pulses + basmati to deepen their 'Indische Spezialitaeten' range; lead with FSSAI/APEDA/Spices Board + better landed pricing vs. EU re-importers"],

[3,"LTP Import Export B.V.","https://www.ltpimpex.com","Netherlands","Hoofddorp",
"B2B Asian-food importer + wholesaler/distributor (pan-European)",
"Rice & Noodles · Flour, Beans & Cereals · Sauces/Pastes/Seasonings · Canned/Dried/Preserved · Beverages",
"2,500+ products across 250+ brands incl. rice, flour, beans & cereals (pulses), seasonings — Asian assortment with a clear South-Asian/Indian whitespace to fill",
"buying@ltpimpex.com · sales@ltpimpex.com","+31 297 309 197","Not publicly available","Not publicly verified",
"https://www.ltpimpex.com/contact","Not applicable (NL — Terms & Privacy only)","https://www.linkedin.com/company/29282937",
"https://www.google.com/maps/search/?api=1&query=LTP%20Import%20Export%20B.V.&query_place_id=ChIJqc3g0tbfxUcRzkkMuNSneVU",
"Active 2026 — 2,500+ products, 250+ brands, ships to 30+ countries; thousands of customer testimonials; dedicated supplier/procurement inbox live",
8,"Enterprise-scale European importer with a DEDICATED procurement inbox (buying@) and existing rice/pulses/seasoning lines but little Indian range = ideal expansion-supplier opportunity for Tbaisa",
"Email buying@ltpimpex.com (their stated supplier/manufacturer contact) proposing an India-origin spice/pulse/basmati range — position as adding a South-Asian category to their 250-brand catalogue; offer private-label option"],

[4,"Oriental Merchant Europe","http://www.orientalmerchant.eu","Netherlands","Oss",
"B2B Asian/oriental food importer + distributor (EU arm of global group founded 1970s)",
"Asian sauces & condiments · pantry staples · multi-brand Asian grocery distribution",
"Distributes 12+ Asian house/partner brands (Chaokoh, Fu Lin Men, Hikari, Prima Taste etc.) into EU retail — currently SE/East-Asian heavy, Indian-origin spices/pulses a clear gap",
"Not publicly available (website contact form only)","+31 88 011 6800","Not publicly available","Not publicly verified",
"https://www.orientalmerchant.eu/contact-us/","Not applicable (NL — Privacy & Terms only)","Not publicly available",
"https://www.google.com/maps/search/?api=1&query=Oriental%20Merchant%20Europe&query_place_id=ChIJ1yAtc-rzxkcRlSVGoT6_Vj8",
"Active 2026 — 2026 catalogue live; global Asian-food specialist operating since 1970s; EU distribution hub in Oss; 4.5 (58 Google reviews)",
7,"Established multi-brand Asian distributor with EU retail reach but no Indian range = strong volume buyer if Tbaisa fills the South-Asian whitespace; only soft point is no public email (contact form)",
"Submit via orientalmerchant.eu/contact-us/ (and ask for the category buyer): pitch becoming their India-origin spice/pulse supplier or private-label brand to add a missing South-Asian line to their EU distribution"],

[5,"Natural Spices B.V.","https://www.naturalspices.nl","Netherlands","Mijdrecht",
"Spice & herb house with B2B 'Zakelijk inkopen' channel (founded 1935)",
"Single herbs & spices · Spice blends (incl. Indian) · Sauces & marinades · Dried vegetables · Organic (Biologisch)",
"Pure single-origin herbs/spices + organic line used to produce blends — sources raw spice material in bulk (incl. Indian-cuisine blends)",
"info@naturalspices.nl","+31 297 254 109","Not publicly available","Not publicly verified",
"https://www.naturalspices.nl/contact","Not applicable (NL — Terms & Privacy only)","https://www.linkedin.com/company/naturalspices",
"https://www.google.com/maps/search/?api=1&query=Natural%20Spices%20%E2%80%93%20Kruiden%20en%20Specerijen&query_place_id=ChIJGaWORq51xkcRGw6UC-HVU9Q",
"Active 2025-2026 — 2025 customer testimonials + new 'Natural Zero' product line; FSSC 22000 certified; ~90 yrs (founded 1935); 134 Google reviews (4.7); explicit B2B portal",
7,"Established spice house with FSSC 22000 + an organic line + a dedicated B2B buying portal = recurring raw-spice buyer for blend production; Indian blends already in range means India-origin demand exists",
"Email info@naturalspices.nl referencing their B2B 'Zakelijk inkopen' channel; offer India-direct single-origin + organic spices for their blend production; emphasise FSSC-compatible specs + ETO-free"],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Germany + NL Leads"

ws.merge_cells("A1:T1")
t = ws.cell(1,1,"Tbaisa Import Export - 5 NEW Ethics-Compliant B2B Buyer Leads (Germany + Netherlands) - found via Apify Google Maps - verified 2026-06-22")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 24

hdr_fill = PatternFill("solid", fgColor="1F4E78")
for c,h in enumerate(headers,1):
    cell = ws.cell(3,c,h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hdr_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[3].height = 30

for r,row in enumerate(leads,4):
    for c,val in enumerate(row,1):
        cell = ws.cell(r,c,val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

widths=[7,30,28,11,16,28,26,40,26,18,16,18,30,30,28,30,42,9,42,48]
for c,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes="A4"

out = r"C:\Users\admin\Downloads\Tbaisa_Germany_Netherlands_5_New_Leads.xlsx"
wb.save(out)
print("SAVED:", out)
