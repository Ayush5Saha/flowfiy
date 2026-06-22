# -*- coding: utf-8 -*-
"""Build the enriched, multi-sheet Tbaisa DE+NL buyer-lead research workbook."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---- Master data (one dict per researched company) ----------------------------
C = [
{
 "no":1,
 "company":"Hermann Schulz GmbH  (Hamburger Gewuerz-Muehle / HGM Spice)",
 "country":"Germany","city":"Hamburg","website":"https://www.hgmspice.de",
 "founded":"1930 (95 yrs)",
 "type":"Spice manufacturer + direct importer + wholesaler + exporter",
 "model":"Direct global import -> in-house milling/cleaning/blending -> wholesale & export",
 "employees":"51-200 (LinkedIn / D&B)",
 "scale":"14 mills, 8 cleaning + 6 mixing lines; ~EUR 10M balance sheet (2023)",
 "entity":"GmbH","reg":"Amtsgericht Hamburg, HRB 30714","vat":"Not publicly listed",
 "parent":"Independent (Schulz family)",
 "dm":"Holger Schulz - Managing Director (GF)",
 "proc":"I. Kelting - Import/Purchasing (Tel -41); D. Thiel - Product Mgmt Organic (Tel -45)",
 "dm_src":"Impressum + company contact page (hgmspice.de/en/contact)",
 "email":"info@hgmspice.de","proc_email":"via Import dept (I. Kelting) contact form",
 "phone":"+49 40 789701-0","contact":"https://www.hgmspice.de/kontakt/",
 "impressum":"https://www.hgmspice.de/kontakt/impressum/",
 "linkedin":"https://www.linkedin.com/company/hgmspice/","social":"Not publicly listed",
 "rating":"4.3 (24 Google reviews)",
 "prod_cat":"Spices A-Z; Blends & Salts; Bio (organic) spices; Specialty items",
 "tbaisa_prod":"Whole & ground raw spices for milling: turmeric, chilli, cumin, coriander, black pepper, ginger; organic (Bio) spices",
 "sourcing":"Global direct import (origin countries worldwide; India a known spice origin)",
 "india_rel":"High - a miller of raw spices is a direct end-buyer of Tbaisa's raw material",
 "gap":"Origin-direct, ETO-free, EU-MRL-compliant raw spice supply + a Bio (organic) origin line",
 "certs_held":"FSSC 22000",
 "certs_need":"EU-MRL + ETO-free residue certs; FSSC-compatible spec sheets; EU-Organic for Bio line",
 "port":"Mundra/Nhava Sheva -> Port of Hamburg (CIF Hamburg)",
 "fit":9,
 "reason":"Direct importer + miller = buys exactly Tbaisa's raw material at volume; named, reachable Import contact; FSSC 22000 forces spec-grade origin supply",
 "priority":"P1 - contact first",
 "channel":"Email (Import dept I. Kelting; cc GF Holger Schulz)",
 "language":"German (DE)",
 "subject":"Indien-Direktbezug: ETO-freie Gewuerze (EU-MRL-konform) fuer Ihre Muehlen",
 "hook":"Reference their 14-mill operation + FSSC 22000; offer India-origin raw turmeric/chilli/cumin/coriander/pepper ex-Mundra, ETO-free & EU-MRL-compliant, with a separate Bio line for D. Thiel's organic range",
 "pitch":"Turmeric (Salem/Erode), Chilli (Teja/Byadgi), Cumin, Coriander, Black Pepper + organic spices",
 "next":"Send spec sheet + EU-MRL/ETO-free CoA samples; request milling-grade volume RFQ",
 "sources":"hgmspice.de; /kontakt/impressum/; northdata.com (HRB 30714); linkedin.com/company/hgmspice; Google Maps",
},
{
 "no":2,
 "company":"Shalimar Foods GmbH",
 "country":"Germany","city":"Cologne (Koeln)","website":"https://shalimar-food.com",
 "founded":"~2001 (24 yrs)",
 "type":"B2B food wholesaler + cash & carry + daily delivery",
 "model":"Importer/wholesaler supplying gastronomy + retail (pickup market + delivery)",
 "employees":"SME (not publicly confirmed)",
 "scale":"1200+ SKUs; cash & carry market + daily delivery fleet",
 "entity":"GmbH","reg":"Amtsgericht Koeln, HRB 38307","vat":"DE 813401933",
 "parent":"Independent",
 "dm":"Ismail Durmaz - Managing Director (GF)",
 "proc":"Ismail Durmaz (owner-operator; central purchasing)",
 "dm_src":"Impressum (shalimar-food.com/impressum)",
 "email":"info@shalimar-foods.com","proc_email":"info@shalimar-foods.com (to GF)",
 "phone":"+49 2203 459690","contact":"https://shalimar-food.com/kontakt/",
 "impressum":"https://shalimar-food.com/impressum/",
 "linkedin":"Not publicly listed","social":"Not publicly listed",
 "rating":"4.8 (24 Google reviews)",
 "prod_cat":"Indian Specialties; Italian/Mexican/Chinese; gastronomy supply",
 "tbaisa_prod":"Indian spices, lentils/pulses, basmati rice, ethnic dry goods",
 "sourcing":"Mixed importers/EU wholesalers (re-import likely); explicit Indian line",
 "india_rel":"High - runs a dedicated 'Indische Spezialitaeten' line for gastronomy",
 "gap":"India-direct pricing & exclusivity vs. EU re-importers; consistent basmati + pulse supply",
 "certs_held":"Not publicly listed (gastronomy supplier; halal likely on request)",
 "certs_need":"FSSAI/APEDA/Spices Board docs; spec sheets; halal optional",
 "port":"Mundra/Nhava Sheva -> Rotterdam/Antwerp -> Cologne (CIF Rotterdam)",
 "fit":9,
 "reason":"South-Asian specialist wholesaler with an explicit Indian line + active distribution (4.8 rating) = natural, ready buyer of India-direct spices/pulses/basmati",
 "priority":"P1 - contact first",
 "channel":"Email to MD Ismail Durmaz",
 "language":"German / English",
 "subject":"India-direct spices, pulses & basmati to widen your Indische Spezialitaeten",
 "hook":"They already run an Indian line + daily delivery; offer direct-from-India landed pricing that beats EU re-importers, backed by FSSAI/APEDA/Spices Board",
 "pitch":"Whole + ground spices, lentils (toor/moong/urad/chana), basmati rice",
 "next":"Offer a starter mixed container + price list; propose sample box to the cash & carry",
 "sources":"shalimar-food.com; /impressum/; Google Maps",
},
{
 "no":3,
 "company":"LTP Import Export B.V.",
 "country":"Netherlands","city":"Hoofddorp","website":"https://www.ltpimpex.com",
 "founded":"2016 (incorporated)",
 "type":"B2B Asian-food importer + wholesaler/distributor (pan-European)",
 "model":"Imports from Asian manufacturers -> distributes to EU supermarkets/wholesalers/retailers",
 "employees":"Mid-size importer (not publicly confirmed)",
 "scale":"2,500+ products; 250+ brands; 10 categories; ships to 30+ EU countries",
 "entity":"B.V.","reg":"KvK 71397930","vat":"NL858699953B01",
 "parent":"Independent",
 "dm":"Not publicly named (check KvK)",
 "proc":"Dedicated procurement inbox: buying@ltpimpex.com",
 "dm_src":"Website (stated supplier/manufacturer contact) + KvK 71397930",
 "email":"sales@ltpimpex.com","proc_email":"buying@ltpimpex.com",
 "phone":"+31 297 309 197","contact":"https://www.ltpimpex.com/contact",
 "impressum":"N/A (NL) - Terms & Privacy only",
 "linkedin":"https://www.linkedin.com/company/29282937",
 "social":"Instagram @ltp_import_export_bv; Facebook /LTPimportexport",
 "rating":"3.4 (10 Google reviews)",
 "prod_cat":"Rice & Noodles; Flour/Beans & Cereals; Sauces/Pastes/Seasonings; Canned/Dried; Beverages",
 "tbaisa_prod":"Rice (basmati), pulses (beans/cereals), spices & seasonings - South-Asian extension",
 "sourcing":"Vietnam, Thailand, China, Taiwan, Korea, Japan, Hong Kong, Indonesia",
 "india_rel":"Medium-High - strong infra & rice/pulse lines but thin Indian/South-Asian range",
 "gap":"A whole India/South-Asian category (spices, pulses, basmati) to add to the 250-brand catalogue",
 "certs_held":"'EU quality control qualified' (self-stated)",
 "certs_need":"EU food-compliance docs; origin (EUR.1); full spec sheets; private-label artwork support",
 "port":"Mundra/Nhava Sheva -> Rotterdam (CIF Rotterdam)",
 "fit":8,
 "reason":"Enterprise-scale European importer with a DEDICATED procurement inbox + existing rice/pulse/seasoning lines but little Indian range = ideal expansion-supplier slot",
 "priority":"P1 - contact first (procurement inbox lowers friction)",
 "channel":"Email buying@ltpimpex.com (their stated supplier contact)",
 "language":"English",
 "subject":"New South-Asian category for your 250-brand range - India-direct",
 "hook":"They cover SE/East Asia but are thin on India; propose adding an India spice/pulse/basmati line or private label to their 2,500-SKU, 30-country operation",
 "pitch":"Spice range, pulses, basmati rice, ready spice blends (private label)",
 "next":"Send catalogue + private-label deck to buying@; propose a trial PO",
 "sources":"ltpimpex.com; creditsafe/D&B; KvK 71397930; linkedin.com/company/29282937; Google Maps",
},
{
 "no":4,
 "company":"Oriental Merchant Europe  (Kaitak B.V.)",
 "country":"Netherlands","city":"Oss","website":"https://www.orientalmerchant.eu",
 "founded":"Group from 1970s (AU); EU arm (Kaitak B.V.) since 2005",
 "type":"B2B Asian-food importer + distributor (EU arm of global Oriental Merchant Group)",
 "model":"Imports authentic Asian brands -> distributes to retailers/wholesalers/C&C across EU",
 "employees":"51-200",
 "scale":"30,000 sqm warehouse, 45,000 pallets; 150+ brands; 1,500+ products; 30+ EU countries",
 "entity":"B.V. (Kaitak B.V., part of Oriental Merchant Group)","reg":"KvK-registered (Oss) - number not publicly confirmed","vat":"Not publicly confirmed",
 "parent":"Oriental Merchant Group (Australia)",
 "dm":"Not publicly named; 'Sales Manager Retail' role identified",
 "proc":"Supplier inquiries via contact form; HR Sabrina van der Mars (hrm@) - not sales",
 "dm_src":"Company site + bestfoodimporters.com profile",
 "email":"Not publicly listed (use website contact form)","proc_email":"via orientalmerchant.eu/contact-us",
 "phone":"+31 88 011 6800","contact":"https://www.orientalmerchant.eu/contact-us/",
 "impressum":"N/A (NL) - Privacy & Terms only",
 "linkedin":"https://www.linkedin.com/company/oriental-merchant-europe",
 "social":"Operates orientalwebshop.nl (Amazing Oriental retail)",
 "rating":"4.5 (58 Google reviews)",
 "prod_cat":"Ethnic Food; Grains; Spices/Herbs/Seasonings; Organic & Health Food; Snacks; Noodles; Tea/Coffee",
 "tbaisa_prod":"Bulk spices/herbs/seasonings, grains/pulses, organic spices - India-origin gap",
 "sourcing":"Predominantly SE/East Asian brands (ABC, Chaokoh, Fu Lin Men, Hikari, Prima Taste...)",
 "india_rel":"Medium-High by scale - explicit Spices/Herbs + Grains + Organic ranges, no Indian origin yet",
 "gap":"India-origin supply behind their Spices/Grains/Organic ranges; private-label India line",
 "certs_held":"In-house QA + Legal/compliance teams (specific certs not listed)",
 "certs_need":"Full QA spec sheets; EU-Organic certs; EU food-compliance; reliable container cadence",
 "port":"Mundra/Nhava Sheva -> Rotterdam (CIF Rotterdam)",
 "fit":8,
 "reason":"Large multi-brand distributor (30k sqm, 30+ countries) listing Spices/Herbs + Grains + Organic but no Indian origin = high-volume slot if Tbaisa fills the whitespace",
 "priority":"P2 - high value but no public sales email (form route)",
 "channel":"Website contact form; ask for category/Sales Manager Retail",
 "language":"English",
 "subject":"Supplier inquiry: India-origin Spices/Herbs & Grains for your EU distribution",
 "hook":"They distribute Spices/Herbs/Seasonings, Grains and Organic to 30+ countries from a 30,000 sqm hub but carry no Indian origin; offer to be that India supplier / private-label source",
 "pitch":"Bulk spices, grains/pulses, organic (EU) spices, private-label India line",
 "next":"Submit supplier inquiry via form; follow up on LinkedIn to reach the category buyer",
 "sources":"orientalmerchant.eu/about-us; bestfoodimporters.com; panjiva.com (buyer report); linkedin.com/company/oriental-merchant-europe; Google Maps",
},
{
 "no":5,
 "company":"Natural Spices B.V.",
 "country":"Netherlands","city":"Mijdrecht","website":"https://www.naturalspices.nl",
 "founded":"1935 (Amsterdam family business); B.V. registered 1998",
 "type":"Spice & herb manufacturer/blender with B2B (wholesale + private label) channel",
 "model":"Buys raw single-origin spices -> blends/produces -> wholesale, foodservice & private label",
 "employees":"~20",
 "scale":"Online groothandel (daily-available bulk) + contract blending + private label",
 "entity":"B.V.","reg":"KvK 36052139 (reg. 26 Jan 1998)","vat":"Not publicly confirmed",
 "parent":"Pecunia non olet B.V. (holding)",
 "dm":"Y.W. Keijlewer - General Director",
 "proc":"B2B / 'Zakelijk inkopen' desk (info@naturalspices.nl)",
 "dm_src":"Drimble/KvK 36052139 + company site",
 "email":"info@naturalspices.nl","proc_email":"info@naturalspices.nl (B2B desk)",
 "phone":"+31 297 254 109","contact":"https://www.naturalspices.nl/contact",
 "impressum":"N/A (NL) - Terms & Privacy only",
 "linkedin":"https://www.linkedin.com/company/naturalspices",
 "social":"Intl B2B site: naturalspices.com/business",
 "rating":"4.7 (134 Google reviews)",
 "prod_cat":"Single herbs & spices; spice blends (incl. Indian); sauces & marinades; organic",
 "tbaisa_prod":"Raw single-origin spices + organic (SKAL) spices as blend inputs",
 "sourcing":"100% natural raw herbs/spices (origins not disclosed) for in-house blending",
 "india_rel":"Medium - blends incl. Indian recipes; needs India single-origin & organic raw material",
 "gap":"India-direct single-origin + EU/SKAL-organic raw spices, ETO-free, for blend production",
 "certs_held":"FSSC 22000; RiskPlaza Audit+; SKAL (organic)",
 "certs_need":"EU/SKAL-organic certs; ETO-free; FSSC-compatible spec sheets; consistent single-origin lots",
 "port":"Mundra/Nhava Sheva -> Rotterdam (CIF Rotterdam)",
 "fit":7,
 "reason":"Established blender/private-label house with FSSC 22000 + SKAL organic + a B2B buying desk = recurring raw-spice buyer; existing Indian blends prove India-origin demand",
 "priority":"P2/P3 - solid recurring buyer, smaller volume",
 "channel":"Email info@ / B2B portal (naturalspices.com/business)",
 "language":"Dutch / English",
 "subject":"India-direct single-origin & SKAL-organic spices for your blends",
 "hook":"They blend + private-label and hold SKAL organic; offer India single-origin + organic raw material, ETO-free and FSSC-compatible, for their recipes",
 "pitch":"Single-origin spices, EU/SKAL-organic-eligible spices, dehydrated ingredients",
 "next":"Email the B2B desk with an organic + conventional single-origin price list; offer samples",
 "sources":"naturalspices.nl + naturalspices.com; drimble.nl (KvK 36052139); linkedin.com/company/naturalspices; Google Maps",
},
]

# ---- Styling helpers ----------------------------------------------------------
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_sheet(ws, title, headers, rows, widths, header_color):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(1, 1, title)
    t.font = Font(bold=True, size=12, color="1F4E78")
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    fill = PatternFill("solid", fgColor=header_color)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    ws.row_dimensions[3].height = 34
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.border = BORDER
        ws.row_dimensions[r].height = 118
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C4"

wb = openpyxl.Workbook()

# Sheet 1 - Company Profile -----------------------------------------------------
ws1 = wb.active
ws1.title = "1) Company Profile"
h1 = ["Lead No.","Company Name","Country","City","Website","Year Founded","Company Type",
      "Business Model","Employees (est.)","Operational Scale","Legal Entity","Registration No.",
      "VAT / Tax ID","Parent / Holding","Primary Decision-Maker","Procurement Contact",
      "DM Source","Official Email","Procurement Email","Phone","Contact Page","Impressum / Legal",
      "LinkedIn","Other Social","Google Rating"]
k1 = ["no","company","country","city","website","founded","type","model","employees","scale",
      "entity","reg","vat","parent","dm","proc","dm_src","email","proc_email","phone","contact",
      "impressum","linkedin","social","rating"]
rows1 = [[c[k] for k in k1] for c in C]
w1 = [7,30,11,16,26,18,26,30,18,30,12,26,18,22,26,30,26,24,26,18,28,30,30,26,16]
style_sheet(ws1, "Tbaisa Import Export - DE + NL Buyer Leads | Sheet 1: COMPANY PROFILE (researched & verified 2026-06-22)", h1, rows1, w1, "1F4E78")

# Sheet 2 - Sourcing & Buyer-Fit ------------------------------------------------
ws2 = wb.create_sheet("2) Sourcing & Buyer-Fit")
h2 = ["Lead No.","Company Name","Product Categories","Relevant Products for Tbaisa",
      "Current Sourcing Regions","India-Relevance","Product Gap / Opportunity",
      "Certifications Held","Certs Tbaisa Must Provide","Shipping Route (Incoterm)",
      "Buyer-Fit Score /10","Reason for Score"]
k2 = ["no","company","prod_cat","tbaisa_prod","sourcing","india_rel","gap","certs_held",
      "certs_need","port","fit","reason"]
rows2 = [[c[k] for k in k2] for c in C]
w2 = [7,30,34,34,30,30,34,26,34,30,12,40]
style_sheet(ws2, "Sheet 2: SOURCING INTELLIGENCE & BUYER-FIT", h2, rows2, w2, "375623")

# Sheet 3 - Outreach Playbook ---------------------------------------------------
ws3 = wb.create_sheet("3) Outreach Playbook")
h3 = ["Lead No.","Company Name","Outreach Priority","Best Channel","Language",
      "Suggested Subject Line","Opening Hook","Products to Pitch","Suggested Next Step",
      "Compliance Line (all leads)"]
compliance = "FSSAI · APEDA · Spices Board India · IEC · BRCGS · FSSC 22000 · EU-MRL · ETO-free · USDA/EU/JAS Organic options"
k3 = ["no","company","priority","channel","language","subject","hook","pitch","next"]
rows3 = [[c[k] for k in k3] + [compliance] for c in C]
w3 = [7,30,26,30,16,40,46,34,38,40]
style_sheet(ws3, "Sheet 3: OUTREACH PLAYBOOK", h3, rows3, w3, "7C4700")

# Sheet 4 - Sources & Ethics ----------------------------------------------------
ws4 = wb.create_sheet("4) Sources & Ethics")
h4 = ["Lead No.","Company Name","Discovery Method","Verification Sources","Data Notes / Ethics"]
ethics = ("All fields from publicly-available business sources (company website, Impressum/KvK, "
          "public directories, Google Business Profile). No logins/CAPTCHA bypassed, no guessed "
          "emails/phones, no private mobiles. 'Not publicly available/confirmed' written honestly where "
          "a field could not be verified.")
rows4 = [[c["no"], c["company"], "Apify - compass/crawler-google-places (Google Maps actor)",
          c["sources"], ethics] for c in C]
w4 = [7,30,40,60,60]
style_sheet(ws4, "Sheet 4: SOURCES & ETHICS-COMPLIANCE LOG", h4, rows4, w4, "5B2C6F")

out = r"C:\Users\admin\Downloads\Tbaisa_Germany_Netherlands_Leads_Researched.xlsx"
wb.save(out)
print("SAVED:", out)
print("Sheets:", wb.sheetnames)
